#!/usr/bin/env python3

import argparse
import sys
from os import path
from typing import Optional, Tuple, List, Dict
from tqdm import tqdm


from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.cell import range_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.cell_range import CellRange


def excel_file(name: str, must_exist: Optional[bool] = None) -> str:
    if not name.lower().endswith('.xlsx'):
        raise ValueError('Not an Excel XLSX file: ' + name)
    if must_exist is not None and path.isfile(name) != must_exist:
        raise ValueError(
            'File ' + ('does not exist' if must_exist else 'already exists') +
            ': ' + name)
    return name


def ws_and_range(wb: Workbook, cell_range: str) -> \
        Tuple[Worksheet, CellRange]:
    sheet_name, boundaries = range_to_tuple(cell_range)
    return wb[sheet_name], CellRange(None, *boundaries)


def header(ws: Worksheet, r: CellRange) -> List[str]:
    it = next(ws.iter_rows(r.min_row, r.min_row, r.min_col, r.max_col, True))
    return list(it)


def main(args: argparse.Namespace) -> None:
    # read file
    in_file = excel_file(args.input_file, must_exist=True)
    out_file = excel_file(
        args.output_file, must_exist=False if not args.overwrite else None)
    tqdm.write('Loading spreadsheet...')
    spreadsheet = load_workbook(in_file, read_only=args.only_new)
    out_spreadsheet = spreadsheet

    # worksheets and ranges
    source_ws, source_range = ws_and_range(spreadsheet, args.source_range)
    key_ws, key_range = ws_and_range(spreadsheet, args.key_range)
    target_ws, target_range = ws_and_range(spreadsheet, args.target_range)

    # ranges must have matching dimensions and must not overlap
    if target_range.size['rows'] != key_range.size['rows']:
        raise Exception('Key and target ranges have different heights')
    if target_ws == key_ws and not target_range.isdisjoint(key_range):
        raise Exception('Key and target ranges overlap')
    if target_ws == source_ws and not target_range.isdisjoint(source_range):
        raise Exception('Source and target ranges overlap')

    # column names
    source_col_names = header(source_ws, source_range)
    key_col_names = header(key_ws, key_range)
    target_col_names = header(target_ws, target_range)
    source_col_names_set = set(source_col_names)
    key_col_names_set = set(key_col_names)
    target_col_names_set = set(target_col_names)

    # if using a fresh sheet for the output, create it now
    blank_left_padding: List[str] = []
    if args.only_new:
        out_spreadsheet = Workbook(write_only=True)
        target_ws = out_spreadsheet.create_sheet(target_ws.title)
        blank_left_padding = (target_range.min_col - 1) * ['', ]
        for _i in range(target_range.min_row - 1):
            target_ws.append([])
        target_ws.append(blank_left_padding + target_col_names)

    # check if some key or target columns are missing from source
    missing = (key_col_names_set | target_col_names_set) - source_col_names_set
    if missing:
        raise Exception('Column(s) missing from source range: ' + str(missing))

    # columns in key range must not be duplicate
    if len(key_col_names_set) < len(key_col_names):
        dup = {c for c in key_col_names_set if key_col_names.count(c) > 1}
        raise Exception('Duplicate column names in key range: ' + str(dup))

    # columns in key or target must be unique in source range
    dup = {c for c in (key_col_names_set | target_col_names_set)
           if source_col_names.count(c) > 1}
    if dup:
        raise Exception('Duplicate column names in source range: ' + str(dup))

    # relevant column indices in source range
    key_in_source = [source_col_names.index(c) for c in key_col_names]
    target_in_source = [source_col_names.index(c) for c in target_col_names]

    # read key columns from source range and save correspondence
    key_to_row_idx: Dict[Tuple, int] = {}
    i0 = source_range.min_row
    all_values: list = []
    for i, row in tqdm(enumerate(source_ws.iter_rows(
        source_range.min_row + 1, source_range.max_row,
        source_range.min_col, source_range.max_col,
        values_only=True)),
            total=source_range.max_row - source_range.min_row,
            desc='Indexing keys', position=0
    ):
        key = tuple(row[j] for j in key_in_source)
        if key in key_to_row_idx:
            old = i0 + key_to_row_idx[key]
            tqdm.write(f'WARNING: Ignoring row {i0 + i + 1} in source range ' +
                       'because their key columns are identical ' +
                       f'to row {old}.')
        else:
            key_to_row_idx[key] = i
        value = tuple(row[j] for j in target_in_source)
        all_values.append(value)

    # read key columns from key range, look it up in source range
    # and fill it in target range
    i0k = key_range.min_row
    i0t, j0t = target_range.min_row, target_range.min_col
    for ik, row in tqdm(enumerate(key_ws.iter_rows(
            key_range.min_row + 1, key_range.max_row,
            key_range.min_col, key_range.max_col,
            values_only=True)),
            total=key_range.max_row - key_range.min_row,
            desc='Filling cells', position=1):
        key = tuple(row[jk] for jk in range(len(key_col_names)))
        try:
            values = all_values[key_to_row_idx[key]]
        except KeyError:
            tqdm.write(f'WARNING: Row {i0k + ik} in key range '
                       + 'has no correspondence in source range.')
            values = len(target_col_names) * (None, )
        else:
            it = ik
            if args.only_new:
                row = [values[jt] for jt in range(len(target_col_names))]
                target_ws.append(blank_left_padding + row)
            else:
                for jt in range(len(target_col_names)):
                    target_ws.cell(i0t + it, j0t + jt, values[jt])
    out_spreadsheet.save(out_file)


if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description='Readd deleted columns from an Excel spreadsheet.')

    parser.add_argument('input_file', type=str)
    parser.add_argument('output_file', type=str)
    parser.add_argument('source_range', type=str)
    parser.add_argument('key_range', type=str)
    parser.add_argument('target_range', type=str)
    parser.add_argument('--overwrite', '-y', action='store_true')
    parser.add_argument('--only-new', action='store_true')
    args = parser.parse_args(sys.argv[1:])

    main(args)
